from openpyxl import load_workbook


def process_fips_codes(filepath):
    """Processes fips codes from xslx file

    Reads fips codes from xslx file and returns dictionary of states and its
    corresponding counties, places and consolidated cities.

    Args:
        filepath (string) (required): Path of the xslx file.
    
    Returns:
        Dictionary of states with corresponding counties, places and 
        consolidated cities.

        Example::
            {
                "01": {
                    "name": "Alabama",
                    "counties": {
                        "01: {
                            "name": "Alabama county",
                            "subdivisions:" {}
                        }
                    },
                    "places": {
                        "01": {
                            "name": "Alabama place"
                        }
                    },
                    "consolidated_cities": {
                        "01": {
                            "name": "Alabama consolidated city"
                        }
                    }
                }
            }
        
    """

    STATE_CODE = 1
    COUNTY_CODE = 2
    COUNTY_SUBDIVISION = 3
    PLACE_CODE = 4
    CONSOLIDATED_CITY = 5
    NAME = 6
    codes = {}
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=6):

        state = row[STATE_CODE].value
        if not state == "00":
            if state not in codes:
                name = row[NAME].value
                codes[state] = {
                    "name": name,
                    "counties": {},
                    "places": {},
                    "consolidated_cities": {}
                }
                continue

            county = row[COUNTY_CODE].value
            if not county == "000":
                if county not in codes[state]["counties"]:
                    name = row[NAME].value
                    codes[state]["counties"][county] = {
                        "name": name,
                        "subdivisions": {}
                    }
                    continue

                county_subdiv = row[COUNTY_SUBDIVISION].value
                if not county_subdiv == "000000":
                    name = row[NAME].value
                    codes[state]["counties"][county]["subdivisions"][
                        county_subdiv] = name
                    continue

            place = row[PLACE_CODE].value
            if not place == "00000":
                name = row[NAME].value
                codes[state]["places"][place] = {"name": name}
                continue

            consolidated_city = row[CONSOLIDATED_CITY].value
            if not consolidated_city == "00000":
                name = row[NAME].value
                codes[state]["consolidated_cities"][consolidated_city] = {
                    "name": name
                }
    return codes


def get_consolidated_cities(states):
    """Prints consolidated cities of states to console.

    Args:
        states (dict) (required): Dictionary of states.
    """
    for state_code, state in states.items():
        for city_code, city in state["consolidated_cities"].items():
            print("city: {} located in state: {}".format(city, state["name"]))


def get_counties_of(state_name, states):
    """Prints counties of state to console.

    Args:
        state_name (string) (required): Name of state to print.
        states (dict) (required): Dictionary of states.
    """
    for state_code, state in states.items():
        if state["name"] == state_name:
            for county_code, county in state["counties"].items():
                print("county: {} of {}".format(county["name"], state_name))

            print("number of counties: {}".format(len(state["counties"])))
            return
